import streamlit as st
import json
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook
import os
import matplotlib.pyplot as plt

# -------------------------
# 🔐 Simple Login
# -------------------------
def check_login():
    user = st.text_input("Username")
    pwd = st.text_input("Password", type="password")
    if user == "admin" and pwd == "iso27001":
        return True
    elif pwd != "":
        st.error("Invalid login")
    return False

if not check_login():
    st.stop()

# -------------------------
# 📥 Load Questions
# -------------------------
try:
    with open("questions.json", "r") as f:
        questions = json.load(f)
except Exception as e:
    st.error(f"Failed to load questions: {e}")
    st.stop()

# -------------------------
# 📝 Audit Setup
# -------------------------
st.title("🛡️ ISO 27001 Internal Audit Tool")
auditor = st.text_input("👤 Enter Auditor Name")
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Session history
if "history" not in st.session_state:
    st.session_state["history"] = []

# Score collection
responses = {}
score_summary = {}
total_score = 0
total_max = 0

# -------------------------
# 📋 Audit Questions
# -------------------------
for domain, qs in questions.items():
    st.subheader(domain)
    domain_score = 0
    for i, q in enumerate(qs):
        answer = st.selectbox(f"{q}", ["Select", "Yes", "No", "Partial"], key=f"{domain}_{i}")
        if answer == "Yes":
            domain_score += 2
        elif answer == "Partial":
            domain_score += 1
    max_score = len(qs) * 2
    score_summary[domain] = {"score": domain_score, "max": max_score}
    total_score += domain_score
    total_max += max_score

# -------------------------
# 📊 Status Helper
# -------------------------
def get_status(score, max_score):
    pct = (score / max_score) * 100
    if pct == 100:
        return "Compliant"
    elif pct >= 50:
        return "Needs Improvement"
    return "Non-Compliant"

# -------------------------
# 📄 Save PDF Report
# -------------------------
def save_pdf(summary, auditor, timestamp):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, "ISO 27001 Audit Report", ln=True, align="C")
    pdf.ln(10)
    pdf.cell(200, 10, f"Auditor: {auditor}", ln=True)
    pdf.cell(200, 10, f"Date: {timestamp}", ln=True)
    pdf.ln(10)
    for domain, data in summary.items():
        status = get_status(data["score"], data["max"])
        pdf.cell(200, 10, f"{domain}: {data['score']} / {data['max']} | Status: {status}", ln=True)
    pdf.ln(10)
    overall_status = get_status(total_score, total_max)
    pdf.cell(200, 10, f"Total Score: {total_score} / {total_max}", ln=True)
    pdf.cell(200, 10, f"Overall Status: {overall_status}", ln=True)
    path = f"audit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(path)
    return path

# -------------------------
# 📊 Save Excel Report
# -------------------------
def save_excel(summary, auditor, timestamp):
    wb = Workbook()
    ws = wb.active
    ws.append(["Auditor", auditor])
    ws.append(["Date", timestamp])
    ws.append(["Domain", "Score", "Max Score"])
    for domain, data in summary.items():
        ws.append([domain, data["score"], data["max"]])
    ws.append(["Total", total_score, total_max])
    path = f"audit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    return path

# -------------------------
# 📈 Chart Visualization
# -------------------------
def show_compliance_chart(summary):
    labels = list(summary.keys())
    scores = [data["score"] for data in summary.values()]
    max_scores = [data["max"] for data in summary.values()]
    compliance_pct = [round((s / m) * 100, 1) for s, m in zip(scores, max_scores)]
    
    st.subheader("📊 Compliance by Domain")
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.barh(labels, compliance_pct, color='skyblue')
    ax.set_xlabel("Compliance %")
    ax.set_title("Audit Compliance Score")
    st.pyplot(fig)

# -------------------------
# 🚀 Generate Reports
# -------------------------
if st.button("Generate Reports"):
    if auditor.strip() == "":
        st.warning("Please enter an auditor name.")
    else:
        pdf_path = save_pdf(score_summary, auditor, timestamp)
        xlsx_path = save_excel(score_summary, auditor, timestamp)
        st.success("Reports generated!")

        # Save session to history
        st.session_state["history"].append({
            "auditor": auditor,
            "timestamp": timestamp,
            "summary": score_summary
        })

        # Download buttons
        st.download_button("📄 Download PDF Report", open(pdf_path, "rb"), file_name=os.path.basename(pdf_path))
        st.download_button("📊 Download Excel Report", open(xlsx_path, "rb"), file_name=os.path.basename(xlsx_path))

        # Chart
        show_compliance_chart(score_summary)

# -------------------------
# 📂 Audit History Display
# -------------------------
if st.session_state["history"]:
    st.subheader("🗂️ Audit History")
    for i, session in enumerate(st.session_state["history"], 1):
        st.markdown(f"**{i}. Auditor:** {session['auditor']} | **Date:** {session['timestamp']}")
        for domain, res in session["summary"].items():
            st.write(f"- {domain}: {res['score']} / {res['max']}")
